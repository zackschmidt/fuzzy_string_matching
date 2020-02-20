import fuzzywuzzy
from fuzzywuzzy import fuzz

import sys
import time
# https://stackoverflow.com/questions/54197145/passing-dynamic-file-path-in-python
from tkinter.filedialog import askopenfilename
from tkinter import Tk
import openpyxl as xl

# Inspiration for project: https://www.datacamp.com/community/tutorials/fuzzy-string-python

# This function calculates the average similarity between the given three cells
# In the case all 3 are empty the function returns 0
def calculate_similarity(cell1, cell2, cell3):
    num_not_empy = 0
    total = 0

    # All three cells empty, this is a problem
    if cell1.value is None and cell2.value is None and cell3.value is None:
        return 0
    # Only one cell has anything in it, flag cell
    elif cell1.value is None and cell3.value == "Unknown":
        return 0
    # Only one cell has anything in it, flag cell
    elif cell2.value is None and cell3.value == "Unknown":
        return 0

    # Calculate cell1 and cell2 similarity if not none
    if cell1.value is not None and cell2.value is not None:
        sim12 = fuzz.token_set_ratio(cell1.value, cell2.value)
        num_not_empy += 1
        total += sim12 
        #print(sim12)
    # Calculate cell2 and cell3 similarity if not none
    if cell2.value is not None and cell3.value is not None and cell3.value != "Unknown":
        sim23 = fuzz.token_set_ratio(cell2.value, cell3.value)
        num_not_empy += 1
        total += sim23
        #print(sim23)
    # Calculate cell3 and cell1 similarity if not none    
    if cell3.value is not None and cell1.value is not None and cell3.value != "Unknown":
        sim31 = fuzz.token_set_ratio(cell3.value, cell1.value)
        num_not_empy += 1
        total += sim31
        #print(sim31)
    # Single cell with no unknwons, flag code
    if total == 0:
        return -1
    # Average between cells with data
    average = total/num_not_empy
    #print(average)
    return average

Tk().withdraw()
filename = askopenfilename()
                    
# open source file
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# calculate total number of rows and columns in source
max_rows = ws1.max_row
max_cols = ws1.max_column

# https://stackoverflow.com/questions/30484220/fill-cells-with-colors-using-openpyxl
# Create colors to fill spreadsheet in with 
red = xl.styles.colors.Color(rgb="00FF0000")
red_fill = xl.styles.fills.PatternFill(patternType="solid", fgColor=red)

green = xl.styles.colors.Color(rgb="008000")
green_fill = xl.styles.fills.PatternFill(patternType="solid", fgColor=green)

yellow = xl.styles.colors.Color(rgb="FFFF00")
yellow_fill = xl.styles.fills.PatternFill(patternType="solid", fgColor=yellow)

# Iterate over all rows in the spreadsheet
for i in range (1, max_rows):
    billing_make = ws1.cell(i+1, 5)
    make = ws1.cell(i+1, 6)
    manuf = ws1.cell(i+1, 7)

    #print(billing_make.value)
    #print(make.value)
    #print(manuf.value)
    tsr1 = calculate_similarity(billing_make, make, manuf)
    #print(tsr1)
    # Cells do not match and there was more than one
    if(tsr1 < 90 and tsr1 != -1):
        billing_make.fill = red_fill
        make.fill = red_fill
        manuf.fill = red_fill
    # Only one cell, flag yellow for attention
    elif(tsr1 == -1):
        billing_make.fill = yellow_fill
        make.fill = yellow_fill
        manuf.fill = yellow_fill
    # Cells have over 90% similarity, they're fine    
    else:
        billing_make.fill = green_fill
        make.fill = green_fill
        manuf.fill = green_fill

    billing_model = ws1.cell(i+1, 8)
    model = ws1.cell(i+1, 9)
    model_id = ws1.cell(i+1, 10)
    #print(billing_model.value)
    #print(model.value)
    #print(model_id.value)
    tsr2 = calculate_similarity(billing_model, model, model_id)
    #print(tsr2)
    # Cells have less than 70% similarity (necessary threshold to encompass long similar results) and there is more than one
    if(tsr2 < 70 and tsr2 != -1):
        billing_model.fill = red_fill
        model.fill = red_fill
        model_id.fill = red_fill
    # Two results have greater than 70% similarity but model_id is unknown...Barb said this was a green case
    elif(tsr2 >= 70 and billing_model.value is not None and model.value is not None and model_id.value == "Unknown"):
        billing_model.fill = green_fill
        model.fill = green_fill
        model_id.fill = green_fill
    # Only one item exists, flag yellow
    elif(tsr1 == -1):
        billing_model.fill = yellow_fill
        model.fill = yellow_fill
        model_id.fill = yellow_fill        
    # All cells match with greater than 70% similarity!
    else:
        billing_model.fill = green_fill
        model.fill = green_fill
        model_id.fill = green_fill

    sn1 = ws1.cell(i+1, 12)
    sn2 = ws1.cell(i+1, 13)
    sn3 = ws1.cell(i+1, 14)
    #print(sn1.value)
    #print(sn2.value)
    #print(sn3.value)

    # All three serial numbers are the same
    if(sn1.value == sn2.value and sn2.value == sn3.value and sn1.value is not None and sn2.value is not None and sn3.value is not None):
        sn1.fill = green_fill
        sn2.fill = green_fill
        sn3.fill = green_fill
    # Two serial numbers are the same
    elif(sn1.value == sn2.value and sn1.value is not None and sn2.value is not None and sn3.value is None):
        sn1.fill = green_fill
        sn2.fill = green_fill
        sn3.fill = green_fill        
    elif(sn2.value==sn3.value and sn2.value is not None and sn3.value is not None and sn1.value is None):
        sn1.fill = green_fill
        sn2.fill = green_fill
        sn3.fill = green_fill 
    elif(sn3.value == sn1.value and sn3.value is not None and sn1.value is not None and sn2.value is None):
        sn1.fill = green_fill
        sn2.fill = green_fill
        sn3.fill = green_fill
    # No serial numbers match 
    else:
        sn1.fill = red_fill
        sn2.fill = red_fill
        sn3.fill = red_fill
    #break 

wb1.save(str(filename))
sys.exit()
#wb2.save(str(filename2))
